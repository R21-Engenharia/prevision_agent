"""
Fase 4 — Relatório Operacional: PACOTES LIBERADOS PARA FVS

REGRA DE LIBERAÇÃO (exata conforme especificação):
  Um pacote está LIBERADO PARA FVS quando:
    1. TODOS os jobs executivos (name != "CONFERÊNCIA FINAL") estão em 100%
    2. O job "CONFERÊNCIA FINAL" existe e está < 100%

  EXCLUÍDOS:
    - Qualquer job executivo < 100% (mesmo 80%, 95%)
    - CONFERÊNCIA FINAL = 100% (pacote já finalizado)
    - Pacotes sem jobs (hasJobs=False)
    - Pacotes sem "CONFERÊNCIA FINAL" (estrutura inesperada — logged)

Formato do relatório: idêntico ao PDF "Pacote de Trabalho - FVS Maio"
  Coluna A: Pacote de Trabalho / FVS  (modelo FVS + código)
  Coluna B: Pavimentos / Locais       (lista compacta de lotes/pavimentos)

Pipeline:
  1. Coleta jobs via floorsPage → activitiesPage → jobs  (ou usa cache)
  2. Aplica regra de liberação
  3. Faz join com QualityAssociations (por activity.id = QA.itemId)
  4. Agrupa por modelo FVS → compact list de locais
  5. Gera Excel
"""
import io, os, sys, json, re, time
os.environ.setdefault("PYTHONUTF8", "1")
if sys.platform == "win32" and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path
from collections import defaultdict
from datetime import datetime
import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── configuração ─────────────────────────────────────────────────────────────

BASE_DIR  = Path(__file__).parent
DATA_DIR  = BASE_DIR / "data" / "raw"
OUT_DIR   = BASE_DIR / "data" / "reports"
OUT_DIR.mkdir(parents=True, exist_ok=True)

token = os.environ.get("PREVISION_TOKEN")
if not token:
    env_path = BASE_DIR / ".env"
    if env_path.exists():
        for line in env_path.read_text().splitlines():
            if line.startswith("PREVISION_TOKEN="):
                token = line.split("=", 1)[1].strip()
                break
if not token:
    print("ERRO: PREVISION_TOKEN nao encontrado em env nem em .env")
    sys.exit(1)

ENDPOINT = "https://api.prevision.com.br/graphql"
HEADERS  = {
    "Content-Type": "application/json",
    "Accept": "application/json",
    "UserAuthorization": f"token {token}",
}
COMPANY_ID     = 479
PAGE_SIZE_ACT  = 10   # atividades por página (menor por causa dos jobs aninhados)
MAX_RETRIES    = 3
TIMEOUT        = 60

PROJECTS = [
    {
        "id": 10223,
        "name": "CAPE TOWN RESIDENCE",
        "building_prefix": "Torre única - ",
        "jobs_cache": DATA_DIR / "10223_jobs_raw.json",
    },
    {
        "id": 18992,
        "name": "HOLMES RESIDENCE",
        "building_prefix": "Holmes Residence - ",
        "jobs_cache": DATA_DIR / "18992_jobs_raw.json",
    },
]

# ─── cliente GraphQL ──────────────────────────────────────────────────────────

def gql(query: str, timeout: int = TIMEOUT, retries: int = MAX_RETRIES):
    """Executa query GraphQL; retorna body dict ou None em caso de falha."""
    for attempt in range(retries):
        try:
            with httpx.Client(follow_redirects=True, timeout=timeout) as c:
                resp = c.post(ENDPOINT, json={"query": query}, headers=HEADERS)
            if resp.status_code == 200:
                body = resp.json()
                if not body.get("errors"):
                    return body
                errs = body["errors"]
                print(f"    ⚠ GraphQL errors: {[e.get('message','')[:120] for e in errs]}")
                return None
            print(f"    ✗ HTTP {resp.status_code} (attempt {attempt+1}/{retries})")
        except Exception as exc:
            print(f"    ✗ Exception: {exc} (attempt {attempt+1}/{retries})")
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
    return None


# ─── coleta de floors ─────────────────────────────────────────────────────────

def collect_floors(project_id: int) -> tuple[list[dict], list[str]]:
    """
    Coleta todos os floors com cursor Relay.
    Retorna (floors_list, cursors_list).
    cursors_list[i] = endCursor após o floor i (para navegar ao floor i+1).
    """
    floors, cursors = [], []
    after = None
    while True:
        after_clause = f', after: "{after}"' if after else ""
        q = f"""{{ me(id: {COMPANY_ID}) {{
          project(id: {project_id}) {{
            floorsPage(first: 50{after_clause}) {{
              pageInfo {{ hasNextPage endCursor }}
              edges {{ node {{ id name position tag }} }}
            }}
          }}
        }}}}"""
        body = gql(q, timeout=30)
        if not body:
            print("  ✗ Falha ao coletar floors")
            break
        page = body["data"]["me"]["project"]["floorsPage"]
        edges = page["edges"]
        for edge in edges:
            floors.append(edge["node"])
            # O cursor desta posição = endCursor da página se for o último, ou
            # precisamos coletar cursor por cursor (first:1 por floor).
            # Solução: usaremos índice + first:1 approach (como em phase3_collect_activities.py)
        pi = page["pageInfo"]
        after = pi["endCursor"]
        if not pi["hasNextPage"]:
            break

    # Para navegação precisa (first:1 per floor), precisamos dos cursores individuais.
    # Re-coleta com first:1 para obter o cursor de cada floor.
    print(f"  Floors encontrados: {len(floors)} — coletando cursores individuais...")
    cursors = []
    prev_cursor = None
    for i, floor in enumerate(floors):
        after_clause = f', after: "{prev_cursor}"' if prev_cursor else ""
        q = f"""{{ me(id: {COMPANY_ID}) {{
          project(id: {project_id}) {{
            floorsPage(first: 1{after_clause}) {{
              pageInfo {{ endCursor }}
              edges {{ node {{ id }} }}
            }}
          }}
        }}}}"""
        body = gql(q, timeout=20)
        if not body:
            print(f"  ✗ Falha ao coletar cursor do floor {i} ({floor.get('name','')})")
            cursors.append(prev_cursor)  # fallback
            continue
        pc = body["data"]["me"]["project"]["floorsPage"]["pageInfo"]["endCursor"]
        cursors.append(pc)
        prev_cursor = pc
        if (i + 1) % 10 == 0:
            print(f"    cursores: {i+1}/{len(floors)}", end="\r")

    print(f"    cursores: {len(floors)}/{len(floors)} ✓          ")
    return floors, cursors


# ─── coleta de activities + jobs por floor ────────────────────────────────────

def collect_jobs_for_floor(
    project_id: int,
    floor_id: int,
    floor_name: str,
    floor_cursor: str | None,
) -> list[dict]:
    """
    Coleta activities com jobs de um floor específico.
    Retorna lista de dicts: {id, wbsCode, hasJobs, jobs: [...]}
    """
    activities = []
    after_cursor = None

    # Navega ao floor correto via cursor
    floor_clause = f', after: "{floor_cursor}"' if floor_cursor else ""

    while True:
        act_after = f', after: "{after_cursor}"' if after_cursor else ""
        q = f"""{{ me(id: {COMPANY_ID}) {{
          project(id: {project_id}) {{
            floorsPage(first: 1{floor_clause}) {{
              edges {{ node {{
                id
                activitiesPage(first: {PAGE_SIZE_ACT}{act_after}) {{
                  totalCount
                  pageInfo {{ hasNextPage endCursor }}
                  edges {{ node {{
                    id
                    wbsCode
                    hasJobs
                    percentageCompleted
                    jobs {{
                      id
                      name
                      percentageCompleted
                    }}
                  }}}}
                }}
              }}}}
            }}
          }}
        }}}}"""

        body = gql(q)
        if not body:
            print(f"    ✗ Falha no floor {floor_name}")
            break

        node = body["data"]["me"]["project"]["floorsPage"]["edges"]
        if not node:
            break
        node = node[0]["node"]

        # Verifica se chegamos no floor correto
        if str(node["id"]) != str(floor_id):
            # Isso pode acontecer se o cursor aponta para um floor errado
            print(f"    ⚠ Floor ID mismatch: esperado {floor_id}, obtido {node['id']}")

        act_page = node["activitiesPage"]
        for edge in act_page["edges"]:
            a = edge["node"]
            activities.append({
                "id":                   a["id"],
                "wbsCode":              a.get("wbsCode"),
                "hasJobs":              a.get("hasJobs", False),
                "percentageCompleted":  a.get("percentageCompleted", 0),
                "jobs":                 a.get("jobs") or [],
            })

        pi = act_page["pageInfo"]
        if not pi["hasNextPage"]:
            break
        after_cursor = pi["endCursor"]

    return activities


# ─── coleta completa de jobs (todos os floors) ────────────────────────────────

def collect_all_jobs(project: dict) -> list[dict]:
    """
    Coleta todos os jobs de todas as activities de um projeto.
    Usa cache se disponível.
    """
    cache_path = project["jobs_cache"]
    pid        = project["id"]
    pname      = project["name"]

    # Verifica cache.
    #
    # ATENCAO: este cache nunca expirava. Uma vez escrito, o arquivo era
    # reaproveitado para sempre — o jobs_raw.json ficou parado em 13/05/2026
    # (69 dias) enquanto o app calculava "pacotes liberados" em cima dele.
    # PREVISION_FORCE_REFRESH=1 ignora o cache e recoleta do Prevision; a
    # coleta agendada usa isso para o dado nao envelhecer indefinidamente.
    forcar = os.getenv("PREVISION_FORCE_REFRESH", "").strip().lower() in {"1", "true", "sim"}

    if cache_path.exists() and not forcar:
        cached = json.loads(cache_path.read_text(encoding="utf-8"))
        activities = cached.get("activities_list", [])
        coletado = cached.get("collected_at", "?")[:10]
        print(f"  ✓ Cache encontrado: {len(activities)} activities (coletado em {coletado})")
        print("    (defina PREVISION_FORCE_REFRESH=1 para recoletar)")
        return activities

    if forcar and cache_path.exists():
        print("  ↻ PREVISION_FORCE_REFRESH ativo — ignorando cache e recoletando")

    print(f"\n  Coletando jobs — {pname} (project {pid})...")
    floors, cursors = collect_floors(pid)
    print(f"  {len(floors)} floors, {len(cursors)} cursores")

    all_acts = []
    total_with_jobs = 0

    for i, (floor, cursor) in enumerate(zip(floors, cursors)):
        # Para o floor[0], usar None como prev_cursor (floorsPage sem after)
        prev_cursor = cursors[i - 1] if i > 0 else None

        acts = collect_jobs_for_floor(pid, floor["id"], floor["name"], prev_cursor)

        for a in acts:
            a["_floor_name"] = floor.get("name", "")
            a["_floor_id"]   = floor["id"]

        floor_with_jobs = sum(1 for a in acts if a.get("hasJobs"))
        total_with_jobs += floor_with_jobs
        all_acts.extend(acts)

        print(
            f"  Floor [{i+1:2d}/{len(floors)}] {floor.get('name','')[:35]:35s} "
            f"| {len(acts):3d} acts | {floor_with_jobs:3d} com jobs",
            end="\r" if i < len(floors)-1 else "\n"
        )

    print(f"\n  Total: {len(all_acts)} activities | {total_with_jobs} com hasJobs=True")

    # Salva cache
    payload = {
        "project_id":   pid,
        "project_name": pname,
        "collected_at": datetime.now().isoformat(),
        "total":        len(all_acts),
        "total_with_jobs": total_with_jobs,
        "activities_list": all_acts,
    }
    cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  ✓ Salvo em: {cache_path}")
    return all_acts


# ─── regra de liberação para FVS ─────────────────────────────────────────────

def is_liberado_para_fvs(jobs: list[dict]) -> tuple[bool, str]:
    """
    Retorna (liberado: bool, motivo: str).

    Regra:
      - Todos os jobs executivos (name != 'CONFERÊNCIA FINAL') devem estar em 100%
      - O job 'CONFERÊNCIA FINAL' deve existir e estar < 100%
    """
    if not jobs:
        return False, "sem jobs"

    executive_jobs  = [j for j in jobs if j.get("name", "").upper() != "CONFERÊNCIA FINAL"]
    conf_final_jobs = [j for j in jobs if j.get("name", "").upper() == "CONFERÊNCIA FINAL"]

    if not conf_final_jobs:
        return False, "sem CONFERÊNCIA FINAL"

    # Verifica executivos
    if not executive_jobs:
        return False, "sem jobs executivos"

    for j in executive_jobs:
        pct = j.get("percentageCompleted") or 0
        if pct < 100:
            return False, f"executivo '{j.get('name','')}' em {pct:.0f}%"

    # Verifica CONFERÊNCIA FINAL
    conf_pct = conf_final_jobs[0].get("percentageCompleted") or 0
    if conf_pct >= 100:
        return False, f"CONFERÊNCIA FINAL já em {conf_pct:.0f}% (pacote finalizado)"

    return True, f"OK — CF em {conf_pct:.0f}%"


def filter_liberados(activities: list[dict]) -> tuple[list[dict], dict]:
    """
    Filtra activities LIBERADAS PARA FVS.
    Retorna (liberados, stats).
    """
    liberados = []
    stats = {
        "total":             len(activities),
        "has_jobs":          0,
        "liberados":         0,
        "sem_jobs":          0,
        "sem_conf_final":    0,
        "exec_incompleto":   0,
        "conf_final_100":    0,
        "sem_jobs_flag":     0,
    }

    for a in activities:
        if not a.get("hasJobs"):
            stats["sem_jobs_flag"] += 1
            continue

        stats["has_jobs"] += 1
        jobs = a.get("jobs") or []

        ok, motivo = is_liberado_para_fvs(jobs)
        if ok:
            liberados.append(a)
            stats["liberados"] += 1
        elif "sem jobs" in motivo and "sem jobs executivos" not in motivo:
            stats["sem_jobs"] += 1
        elif "sem CONFERÊNCIA FINAL" in motivo:
            stats["sem_conf_final"] += 1
        elif "CONFERÊNCIA FINAL já" in motivo:
            stats["conf_final_100"] += 1
        else:
            stats["exec_incompleto"] += 1

    return liberados, stats


# ─── parse de locais (mesma lógica do relatorio v1) ──────────────────────────

FLOOR_ABBREV = {
    "pavimento tipo": "PV",
    "pavimento diferenciado": "PV Dif.",
    "diferenciado": "Dif.",
    "lazer": "Lazer",
    "platibanda": "Platibanda",
    "barrilete": "Barrilete",
    "reservatório": "Reservatório",
    "reservatorio": "Reservatório",
    "cobertura": "Cobertura",
    "térreo": "Térreo",
    "terreo": "Térreo",
    "estacionamento público": "Est. Público",
    "estacionamento publico": "Est. Público",
    "mezanino": "Mez.",
    "subsolo": "Sub.",
    "casa de máquinas": "CMQ",
    "casa de maquinas": "CMQ",
    "pós grua": "Pós Grua",
    "pos grua": "Pós Grua",
    "pós cremalheira": "Pós Cremalheira",
    "pos cremalheira": "Pós Cremalheira",
    "teto do reservatório": "Teto Reserv.",
    "teto do reservatorio": "Teto Reserv.",
    "salas comerciais": "Salas Com.",
}

SPACE_ABBREV = {
    "apartamento": "Apto",
    "ambientes comuns internos": "AC Int.",
    "ambientes comuns externos": "AC Ext.",
    "ambientes comuns ( rooftop )": "AC Rooftop",
    "ambientes comuns": "AC",
    "área de circulação de veículos": "Circ. Veíc.",
    "area de circulacao de veiculos": "Circ. Veíc.",
    "antecâmara": "Antec.",
    "antecamara": "Antec.",
    "hall": "Hall",
    "escadaria": "Escad.",
    "lajes": "Lajes",
    "pilares": "Pilares",
    "vigas": "Vigas",
    "estrutura": "Estrutura",
    "modificações clientes": "Modif. Clientes",
    "modificacoes clientes": "Modif. Clientes",
}


def parse_floor_short(floor_desc: str) -> tuple[int | None, str]:
    d  = floor_desc.strip()
    dl = d.lower()

    m = re.match(r'^(\d+)[°º]', d)
    floor_num = int(m.group(1)) if m else None

    if m and "pavimento tipo" in dl:
        return floor_num, f"{m.group(1)}º PV"
    if m and "pavimento diferenciado" in dl:
        rest = re.sub(r"^\d+[°º]\s*", "", d, flags=re.IGNORECASE)
        rest = re.sub(r"pavimento diferenciado", "PV Dif.", rest, flags=re.IGNORECASE)
        return floor_num, f"{m.group(1)}º {rest.strip()}"
    if m and "diferenciado" in dl:
        rest = re.sub(r"^\d+[°º]\s*", "", d, flags=re.IGNORECASE)
        return floor_num, f"{m.group(1)}º {rest.strip()}"
    if re.match(r'^G\d', d, re.IGNORECASE):
        return None, d
    if "casa de m" in dl:
        num = re.search(r'\d+', d)
        n   = num.group(0) if num else ""
        return None, f"CMQ {n}".strip()
    for key, abbr in FLOOR_ABBREV.items():
        if key in dl:
            return (0 if "térreo" in dl or "terreo" in dl else None), abbr
    return floor_num, d


def parse_space_short(space_desc: str) -> str:
    d  = space_desc.strip()
    dl = d.lower()
    m  = re.match(r'apartamento\s+(\d+)', dl)
    if m:
        return m.group(1)
    for key, abbr in SPACE_ABBREV.items():
        if dl == key or dl.startswith(key + " ") or dl.startswith(key + "\n"):
            return abbr
    return d


def parse_local(local_full: str, building_prefix: str) -> tuple[int | None, str, str | None]:
    local = local_full.strip()
    if local.startswith(building_prefix):
        local = local[len(building_prefix):]
    parts = [p.strip() for p in local.split(" - ")]
    if not parts:
        return None, local_full, None
    floor_num, floor_short = parse_floor_short(parts[0])
    if len(parts) >= 2:
        space = parse_space_short(parts[1])
        if len(parts) >= 3:
            sub   = parse_space_short(parts[2])
            space = f"{space} / {sub}"
    else:
        space = None
    return floor_num, floor_short, space


def compact_locais(locais_list: list[tuple]) -> str:
    by_floor: dict[str, dict]    = {}
    floor_order: dict[str, int]  = {}

    for floor_num, floor_short, space in locais_list:
        key = floor_short
        if key not in by_floor:
            by_floor[key]    = {"floor_num": floor_num, "spaces": set()}
            floor_order[key] = floor_num if floor_num is not None else 9999
        if space:
            by_floor[key]["spaces"].add(space)

    sorted_floors = sorted(by_floor.keys(), key=lambda k: floor_order.get(k, 9999))

    def fmt_one(floor_short: str, spaces: set) -> str:
        if not spaces:
            return floor_short
        sorted_spaces = sorted(spaces, key=lambda s: (not s.isdigit(), s))
        return f"{floor_short} ({', '.join(sorted_spaces)})"

    result_parts = []
    i = 0
    while i < len(sorted_floors):
        fs = sorted_floors[i]
        fn = by_floor[fs]["floor_num"]
        sp = by_floor[fs]["spaces"]
        is_tipo = fs.endswith(" PV") or fs == "PV"

        j = i + 1
        if is_tipo and fn is not None:
            while j < len(sorted_floors):
                fs2     = sorted_floors[j]
                fn2     = by_floor[fs2]["floor_num"]
                sp2     = by_floor[fs2]["spaces"]
                is_t2   = fs2.endswith(" PV") or fs2 == "PV"
                if is_t2 and fn2 is not None and fn2 == fn + (j - i) and sp2 == sp:
                    j += 1
                else:
                    break

        if j - i >= 3:
            fs_start = sorted_floors[i]
            fs_end   = sorted_floors[j - 1]
            ms       = re.match(r'^(\d+)º', fs_start)
            me_      = re.match(r'^(\d+)º', fs_end)
            if ms and me_:
                range_str = f"{ms.group(1)}º ao {me_.group(1)}º PV"
            else:
                range_str = f"{fs_start} ao {fs_end}"
            if sp:
                sorted_spaces = sorted(sp, key=lambda s: (not s.isdigit(), s))
                range_str += f" ({', '.join(sorted_spaces)})"
            result_parts.append(range_str)
            i = j
        else:
            for k2 in range(i, j):
                result_parts.append(fmt_one(sorted_floors[k2], by_floor[sorted_floors[k2]]["spaces"]))
            i = j

    return ", ".join(result_parts)


# ─── extração FVS ─────────────────────────────────────────────────────────────

def extract_fvs_code(model_name: str) -> tuple[str, str]:
    m = re.match(r'FVS\s+([\d.]+)\s*[-–]\s*(.+)', model_name.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "", model_name.strip()


def format_fvs_label(model_name: str) -> str:
    code, desc = extract_fvs_code(model_name)
    if code:
        return f"{desc} (FVS {code})"
    return model_name


# ─── processamento principal por projeto ──────────────────────────────────────

def process_project(project: dict, activities: list[dict]) -> tuple[list[dict], dict]:
    """
    Filtra activities liberadas, faz join com QAs, gera linhas do relatório.
    Retorna (rows, stats).
    """
    pid    = project["id"]
    prefix = project["building_prefix"]

    # Carrega QAs
    qa_file = DATA_DIR / f"{pid}_qa_raw.json"
    qa_raw  = json.loads(qa_file.read_text(encoding="utf-8"))
    qas     = qa_raw.get("quality_associations", qa_raw if isinstance(qa_raw, list) else [])
    print(f"  QAs carregadas: {len(qas)}")

    # Aplica regra de liberação
    liberados, stats = filter_liberados(activities)
    print(f"  Liberados para FVS: {stats['liberados']} / {stats['has_jobs']} (com jobs)")
    print(f"    exec incompleto:   {stats['exec_incompleto']}")
    print(f"    cf já finalizada:  {stats['conf_final_100']}")
    print(f"    sem conf final:    {stats['sem_conf_final']}")
    print(f"    sem jobs flag:     {stats['sem_jobs_flag']}")

    # Índice de activities liberadas por id
    liberados_ids = {str(a["id"]) for a in liberados}

    # Constrói QA index: activity_id → list de QAs
    qa_by_activity: dict[str, list] = defaultdict(list)
    for qa in qas:
        aid = str(qa.get("itemId", ""))
        if aid in liberados_ids:
            qa_by_activity[aid].append(qa)

    activities_com_qa = len([aid for aid in liberados_ids if qa_by_activity.get(aid)])
    print(f"  Atividades liberadas com QA: {activities_com_qa} / {len(liberados_ids)}")

    # Agrupa por modelo FVS → locais
    model_locais:   dict[str, list]  = defaultdict(list)
    model_act_ids:  dict[str, set]   = defaultdict(set)

    sem_qa_count = 0
    for a in liberados:
        aid = str(a["id"])
        act_qas = qa_by_activity.get(aid)
        if not act_qas:
            sem_qa_count += 1
            continue

        for qa in act_qas:
            ref   = qa.get("partnerReferenceName", "")
            parts = ref.split(" | ", 1)
            if len(parts) < 2:
                continue
            model_name = parts[0].strip()
            local_full = parts[1].strip()
            fn, fs, sp = parse_local(local_full, prefix)
            model_locais[model_name].append((fn, fs, sp))
            model_act_ids[model_name].add(aid)

    if sem_qa_count:
        print(f"  ⚠ Atividades liberadas sem QA (hasQualityAssociations=False?): {sem_qa_count}")

    # Gera linhas
    rows = []
    for model_name, locais in model_locais.items():
        code, _     = extract_fvs_code(model_name)
        label       = format_fvs_label(model_name)
        locais_str  = compact_locais(locais)
        rows.append({
            "fvs_code":  code or "99.99.99",
            "fvs_label": label,
            "model_raw": model_name,
            "locais_str": locais_str,
            "count":     len(locais),
            "n_acts":    len(model_act_ids[model_name]),
        })

    # Ordena por código FVS
    def sort_key(r):
        try:
            return tuple(int(p) for p in r["fvs_code"].replace(".", " ").split())
        except Exception:
            return (99, 99, 99)

    rows.sort(key=sort_key)
    print(f"  Modelos FVS a reportar: {len(rows)}")
    return rows, stats


# ─── geração do Excel ─────────────────────────────────────────────────────────

COLOR_HEADER_BG  = "2F5597"
COLOR_HEADER_FG  = "FFFFFF"
COLOR_ROW_ODD    = "FFFFFF"
COLOR_ROW_EVEN   = "EEF2F8"
COLOR_BORDER     = "BFBFBF"
COLOR_TITLE_BG   = "1F3864"


def make_border(style="thin"):
    s = Side(border_style=style, color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def write_report_sheet(ws, rows: list[dict], project_name: str, stats: dict):
    total_libs = stats.get("liberados", 0)
    total_acts = stats.get("total", 0)

    # ── Título
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = f"PACOTE DE TRABALHO / FVS  —  {project_name}  |  PACOTES LIBERADOS PARA FVS"
    c.font      = Font(bold=True, color=COLOR_HEADER_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    # ── Subtítulo / meta
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = (
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
        f"{total_libs} pacotes liberados de {total_acts} total  |  "
        f"{len(rows)} modelos FVS"
    )
    c.font      = Font(italic=True, color="666666", size=9)
    c.fill      = PatternFill("solid", fgColor="F2F2F2")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Cabeçalho
    hdr_row = 3
    for col, (header, width) in enumerate([
        ("Pacote de Trabalho / FVS", 42),
        ("Pavimentos / Locais Associados", 70),
    ], start=1):
        cell = ws.cell(row=hdr_row, column=col, value=header)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[hdr_row].height = 20

    # ── Dados
    for i, row in enumerate(rows, start=1):
        excel_row = hdr_row + i
        bg_color  = COLOR_ROW_ODD if i % 2 == 1 else COLOR_ROW_EVEN

        ca = ws.cell(row=excel_row, column=1, value=row["fvs_label"])
        ca.font      = Font(bold=True, size=9)
        ca.fill      = PatternFill("solid", fgColor=bg_color)
        ca.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ca.border    = make_border()

        cb = ws.cell(row=excel_row, column=2, value=row["locais_str"])
        cb.font      = Font(size=9)
        cb.fill      = PatternFill("solid", fgColor=bg_color)
        cb.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cb.border    = make_border()

        approx_lines = max(1, len(row["locais_str"]) // 80 + 1)
        ws.row_dimensions[excel_row].height = max(18, approx_lines * 14)

    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)


def generate_excel(project_data: list[tuple]) -> Path:
    """
    project_data: list de (project_dict, rows, stats)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for project, rows, stats in project_data:
        sheet_name = project["name"][:31]
        ws = wb.create_sheet(title=sheet_name)
        write_report_sheet(ws, rows, project["name"], stats)
        print(f"  ✓ Aba '{sheet_name}': {len(rows)} modelos FVS")

    date_str  = datetime.now().strftime("%Y%m%d_%H%M")
    filename  = f"FVS_Liberados_{date_str}.xlsx"
    out_path  = OUT_DIR / filename
    wb.save(out_path)
    return out_path


# ─── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("═" * 65)
    print("  FASE 4 — Relatório FVS Liberados (Regra Exata)")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("═" * 65)
    print()
    print("  REGRA:")
    print("    LIBERADO = todos exec_jobs em 100%  AND  CONF_FINAL < 100%")
    print()

    project_data = []
    for proj in PROJECTS:
        print(f"\n{'─'*65}")
        print(f"  PROJETO: {proj['name']} (id: {proj['id']})")
        print(f"{'─'*65}")

        t0 = time.time()
        activities = collect_all_jobs(proj)
        elapsed    = time.time() - t0
        print(f"  Coleta: {elapsed:.1f}s")

        rows, stats = process_project(proj, activities)
        project_data.append((proj, rows, stats))

    print(f"\n{'═'*65}")
    print("  Gerando Excel...")
    out_path = generate_excel(project_data)
    print(f"\n  ✓ Salvo em: {out_path.resolve()}")
    print("═" * 65)
    print("  CONCLUÍDO")
    print("═" * 65)
