"""
Relatorio Excel do Backlog FVS — respeita os filtros aplicados na tela.

Reaproveita os helpers de formatacao de fvs_dashboard/core/exporter.py.
"""
from __future__ import annotations

import datetime
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from fvs_dashboard.core.exporter import (
    STATUS_COLORS, STATUS_LABEL, _border, _fill, _font, _set_w, _write_row,
)

# Paleta R21 (mesma identidade do app web)
C_VERMELHO_R21 = "C41230"
C_CINZA_ESC    = "3B3F44"
BG_CINZA_CLARO = "F0F2F5"

COLUNAS = [
    ("Pavimento",   26, "left"),
    ("WBS",         11, "center"),
    ("CF%",          8, "center"),
    ("Modelo FVS",  46, "left"),
    ("Local",       34, "left"),
    ("Status",      16, "center"),
    ("% Exec",       9, "center"),
    ("NC",           7, "center"),
    ("NC Pend.",     9, "center"),
    ("Inspecao",    13, "center"),
]


def _titulo(ws, texto: str, ncols: int, row: int, cor: str = C_VERMELHO_R21,
            tamanho: int = 13, altura: int = 26) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row, 1, texto)
    c.fill = _fill(cor)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=tamanho)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = altura


def _linha_info(ws, texto: str, ncols: int, row: int) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row, 1, texto)
    c.fill = _fill(BG_CINZA_CLARO)
    c.font = Font(name="Calibri", size=9, color=C_CINZA_ESC)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18


def build_backlog_report(
    rows: list[dict[str, Any]],
    obra: str,
    descricao_filtros: str,
    total_geral: int,
) -> bytes:
    """Gera um .xlsx com as FVS ja filtradas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog FVS"

    ncols = len(COLUNAS)
    gerado = datetime.datetime.now().strftime("%d/%m/%Y as %H:%M")

    # ── Cabecalho ─────────────────────────────────────────────────────────────
    _titulo(ws, f"BACKLOG FVS — {obra.upper()}", ncols, 1)
    _linha_info(ws, f"Filtros aplicados: {descricao_filtros}", ncols, 2)
    _linha_info(ws, f"{len(rows)} de {total_geral} FVS  |  Gerado em {gerado}", ncols, 3)

    # ── Resumo por status ─────────────────────────────────────────────────────
    contagem = {
        "FINALIZADA":   sum(1 for r in rows if r["status"] == "FINALIZADA"),
        "EM_ANDAMENTO": sum(1 for r in rows if r["status"] == "EM_ANDAMENTO"),
        "NAO_INICIADA": sum(1 for r in rows if r["status"] == "NAO_INICIADA"),
    }
    nc_abertas = sum(int(r.get("nc") or 0) for r in rows)

    resumo_row = 5
    resumo = [
        ("Finalizadas",   contagem["FINALIZADA"],   "FINALIZADA"),
        ("Em andamento",  contagem["EM_ANDAMENTO"], "EM_ANDAMENTO"),
        ("Nao iniciadas", contagem["NAO_INICIADA"], "NAO_INICIADA"),
        ("NC abertas",    nc_abertas,               None),
    ]
    col = 1
    for rotulo, valor, status in resumo:
        fg, bg = STATUS_COLORS.get(status, (C_CINZA_ESC, BG_CINZA_CLARO))
        ws.merge_cells(start_row=resumo_row, start_column=col,
                       end_row=resumo_row, end_column=col + 1)
        c = ws.cell(resumo_row, col, f"{rotulo}: {valor}")
        c.fill = _fill(bg)
        c.font = Font(name="Calibri", bold=True, size=10, color=fg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
        col += 2
    ws.row_dimensions[resumo_row].height = 20

    # ── Cabecalho da tabela ───────────────────────────────────────────────────
    head_row = 7
    for j, (nome, largura, _al) in enumerate(COLUNAS, 1):
        c = ws.cell(head_row, j, nome)
        c.fill = _fill(C_CINZA_ESC)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
        _set_w(ws, j, largura)
    ws.row_dimensions[head_row].height = 20

    # ── Linhas ────────────────────────────────────────────────────────────────
    for i, r in enumerate(rows):
        linha = head_row + 1 + i
        status = r.get("status", "")
        fg, bg = STATUS_COLORS.get(status, (C_CINZA_ESC, "FFFFFF"))

        pct = r.get("pct_exec")
        valores = [
            r.get("floor", ""),
            r.get("wbs", ""),
            f"{float(r.get('cf_pct') or 0):.0f}%",
            r.get("modelo", ""),
            r.get("local", ""),
            STATUS_LABEL.get(status, status),
            "—" if pct is None else f"{pct}%",
            int(r.get("nc") or 0) or "",
            int(r.get("nc_pendentes") or 0) or "",
            r.get("data_ins", "") or "—",
        ]
        fills = ["FFFFFF"] * len(COLUNAS)
        fills[5] = bg                                  # coluna Status
        fonts = [_font(size=9) for _ in COLUNAS]
        fonts[5] = _font(color=fg, bold=True, size=9)
        aligns = [al for _n, _w, al in COLUNAS]

        _write_row(ws, linha, valores, fills, fonts, aligns)

        # Link do InMeta na coluna Modelo FVS
        link = r.get("link") or ""
        if link:
            c = ws.cell(linha, 4)
            c.hyperlink = link
            c.font = _font(color="0563C1", size=9)

    # ── Congelar cabecalho + autofiltro ───────────────────────────────────────
    ws.freeze_panes = ws.cell(head_row + 1, 1)
    ws.auto_filter.ref = (
        f"A{head_row}:{get_column_letter(ncols)}{head_row + len(rows)}"
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
