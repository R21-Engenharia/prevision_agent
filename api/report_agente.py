"""
Relatório Excel das Pendências Inteligentes.
============================================
Modelo que também servirá de base para os e-mails (Fase 4). Duas naturezas:
  tipo="obra" → atrasos que travam o cronograma (foco: impacto)
  tipo="fvs"  → serviços ~prontos, falta a ficha (foco: o que preencher)
"""
from __future__ import annotations

import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_ACCENT = "C41230"
_OK = "2F9E6A"
_INK = "12151A"
_HAIR = "DADEE3"

_thin = Side(style="thin", color=_HAIR)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _jobs(causa: dict | None) -> str:
    jobs = (causa or {}).get("jobs_pendentes") or []
    return "; ".join(f"{j.get('name', '')} ({j.get('pct', 0):.0f}%)" for j in jobs[:4])


def build_pendencias_report(obra: str, tipo: str, pendencias: list[dict]) -> bytes:
    e_obra = tipo == "obra"
    titulo = "Atrasos de obra" if e_obra else "Pendências de FVS"
    cor = _ACCENT if e_obra else _OK

    wb = Workbook()
    ws = wb.active
    ws.title = "Pendências"

    # Cabeçalho do relatório
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"R21 · Agente do Planejamento — {titulo}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=cor)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    hoje = datetime.date.today().strftime("%d/%m/%Y")
    sub.value = f"Obra: {obra}   ·   Gerado em {hoje}   ·   {len(pendencias)} pendências"
    sub.font = Font(size=10, color="59616B")
    sub.alignment = Alignment(horizontal="left", indent=1)

    # Colunas conforme a natureza
    if e_obra:
        cols = ["Pacote", "Pavimento", "Avanço", "Trava (serviços)", "Serviço pendente", "Categoria"]
    else:
        cols = ["Pacote", "Pavimento", "Avanço", "Falta preencher", "Serviço pendente", "Status"]

    hrow = 4
    for j, nome in enumerate(cols, start=1):
        cell = ws.cell(row=hrow, column=j, value=nome)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_INK)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = _BORDER
    ws.row_dimensions[hrow].height = 20

    # Ordena por impacto (obra) ou pavimento (fvs)
    if e_obra:
        pend = sorted(pendencias, key=lambda p: -(p.get("impacto") or 0))
    else:
        pend = sorted(pendencias, key=lambda p: (p.get("pavimento") or ""))

    r = hrow + 1
    for p in pend:
        causa = p.get("causa_raiz") or {}
        jobs = _jobs(causa)
        avanco = f"{(p.get('pct_real') or 0):.0f}%"
        if e_obra:
            linha = [p.get("servico", ""), p.get("pavimento", ""), avanco,
                     p.get("impacto") or 0, jobs, p.get("categoria", "")]
        else:
            linha = [p.get("servico", ""), p.get("pavimento", ""), avanco,
                     "FVS / conferência", jobs, p.get("status", "")]
        for j, val in enumerate(linha, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(size=10, color=_INK)
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            cell.border = _BORDER
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F6F7F9")
        r += 1

    # Larguras
    larguras = [34, 24, 9, 15, 40, 16]
    for j, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{hrow}:F{max(hrow, r - 1)}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
