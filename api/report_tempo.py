"""
Exportacao Excel da Condicao do Tempo — segue o padrao do arquivo
"Diario do Tempo - Obras - Modelo.xlsx" usado nas reunioes.

Padrao visual extraido do modelo:
  - aba RESUMO com titulos em negrito e pizzas "Condicao do Tempo" lado a lado
  - abas DADOS com as colunas Data | Obra | Classificacao do tempo |
    Condicao de trabalho, e um bloco de resumo nas colunas G..L
  - cores das fatias: Ensolarado #FFC000, Nublado #A6A6A6, Chuvoso #3399FF
  - rotulos de dados exibindo percentual
  - formulas vivas (COUNTIF/SUM), como no arquivo original

Diferenca proposital: no modelo a coluna K somava a contagem com a coluna de
percentual (K=H+I+J com I sendo H/$H$30), o que inflava o total do periodo.
Aqui o total e a contagem pura — o visual e o mesmo, o numero e o correto.
"""
from __future__ import annotations

import datetime
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.data_source import StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Cores do modelo
COR = {
    "Ensolarado": "FFC000",
    "Nublado":    "A6A6A6",
    "Chuvoso":    "3399FF",
}
CATEGORIAS = ["Ensolarado", "Nublado", "Chuvoso"]

# Simbolos usados nas reunioes. Precisam aparecer TANTO na coluna C (dados)
# quanto na coluna G (resumo): o COUNTIF casa os dois textos, entao decorar
# so um dos lados zeraria todas as contagens.
SIMBOLO = {
    "Ensolarado": "☀️",
    "Nublado":    "⛅",
    "Chuvoso":    "🌧️",
}


def rotulo(categoria: str) -> str:
    """'Ensolarado' -> '☀️ Ensolarado'"""
    simbolo = SIMBOLO.get(categoria, "")
    return f"{simbolo} {categoria}".strip()

_MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
          "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

_BORDA = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _cabecalho_tabela(ws, linha: int = 1) -> None:
    titulos = ["Data", "Obra", "Classificação do tempo", "Condição de trabalho"]
    larguras = [20, 26, 24, 22]
    for i, (t, w) in enumerate(zip(titulos, larguras), start=1):
        c = ws.cell(linha, i, t)
        c.font = Font(name="Calibri", bold=True, size=11)
        c.border = _BORDA
        ws.column_dimensions[get_column_letter(i)].width = w


def _bloco_resumo(
    ws,
    linha_ini: int,
    titulo_bloco: str,
    historico: dict[str, int] | None,
) -> tuple[int, int]:
    """
    Escreve o bloco G..L no padrao do modelo.
    Retorna (linha_da_primeira_categoria, linha_do_total).
    """
    ws.cell(linha_ini, 9, titulo_bloco).font = Font(name="Calibri", bold=True, size=11)

    linha_total = linha_ini + 1
    l0 = linha_ini + 2                      # primeira categoria

    usa_hist = bool(historico)
    col_total = 11 if usa_hist else 8       # K quando ha historico, senao H

    for i, cat in enumerate(CATEGORIAS):
        r = l0 + i
        ws.cell(r, 7, rotulo(cat)).font = Font(name="Calibri", size=11)
        # Contagem do que veio do InMeta (formula viva, como no modelo)
        ws.cell(r, 8, f"=COUNTIF(C:C,G{r})")
        if usa_hist:
            ws.cell(r, 9, int(historico.get(cat, 0)))          # historico interno
            ws.cell(r, 11, f"=H{r}+I{r}")                      # total
        letra_total = get_column_letter(col_total)
        c_pct = ws.cell(r, 12, f"={letra_total}{r}/${letra_total}${linha_total}")
        c_pct.number_format = "0.0%"

    letra = get_column_letter(col_total)
    c_tot = ws.cell(linha_total, col_total, f"=SUM({letra}{l0}:{letra}{l0 + 2})")
    c_tot.font = Font(name="Calibri", bold=True, size=11)

    # Rotulos das colunas do bloco
    ws.cell(l0 - 1, 8, "InMeta").font = Font(size=9, color="808080")
    if usa_hist:
        ws.cell(l0 - 1, 9, "Histórico").font = Font(size=9, color="808080")
        ws.cell(l0 - 1, 11, "Total").font = Font(size=9, color="808080")
    ws.cell(l0 - 1, 12, "%").font = Font(size=9, color="808080")

    for col in range(7, 13):
        ws.column_dimensions[get_column_letter(col)].width = 14

    return l0, col_total


def _pizza(ws_dados, titulo: str, l0: int, col_total: int) -> PieChart:
    """Pizza no padrao do modelo: cores fixas por categoria e rotulo em %."""
    ch = PieChart()
    ch.title = titulo
    ch.height = 7.5
    ch.width = 15

    dados = Reference(ws_dados, min_col=col_total, min_row=l0, max_row=l0 + 2)
    cats = Reference(ws_dados, min_col=7, min_row=l0, max_row=l0 + 2)
    ch.add_data(dados, titles_from_data=False)
    ch.set_categories(cats)

    serie = ch.series[0]
    serie.data_points = [
        DataPoint(idx=i, spPr=None) for i in range(len(CATEGORIAS))
    ]
    for i, cat in enumerate(CATEGORIAS):
        ponto = serie.data_points[i]
        ponto.graphicalProperties.solidFill = COR[cat]
        ponto.graphicalProperties.line.solidFill = "FFFFFF"
        ponto.graphicalProperties.line.width = 19050   # 1.5pt

    # Rotulo dentro da fatia: nome da categoria + percentual.
    # showSerName precisa ser explicitamente False — sem isso o Excel escreve
    # "Serie1; 28%" no lugar do nome da condicao.
    ch.dataLabels = DataLabelList()
    ch.dataLabels.showPercent = True
    ch.dataLabels.showCatName = True
    ch.dataLabels.showSerName = False
    ch.dataLabels.showVal = False
    ch.dataLabels.showLegendKey = False
    ch.dataLabels.showBubbleSize = False
    ch.dataLabels.dLblPos = "inEnd"
    return ch


_BORDA_CAIXA = Border(*[Side(style="medium", color="000000")] * 4)


def _caixa(ws, celula_ini: str, celula_fim: str, texto: str, tamanho: int = 11) -> None:
    """Caixa com borda e texto centralizado, como no modelo das reunioes."""
    ws.merge_cells(f"{celula_ini}:{celula_fim}")
    c = ws[celula_ini]
    c.value = texto
    c.font = Font(name="Calibri", bold=True, size=tamanho)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    col_ini = ws[celula_ini].column
    col_fim = ws[celula_fim].column
    for r in range(ws[celula_ini].row, ws[celula_fim].row + 1):
        for col in range(col_ini, col_fim + 1):
            ws.cell(r, col).border = _BORDA_CAIXA


def _escreve_aba_dados(
    wb: Workbook,
    nome: str,
    dias: list[dict[str, Any]],
    rotulo_bloco: str,
    historico: dict[str, int] | None,
) -> tuple[Any, int, int, int]:
    ws = wb.create_sheet(nome[:31])
    _cabecalho_tabela(ws)

    for i, d in enumerate(dias, start=2):
        ws.cell(i, 1, datetime.date.fromisoformat(d["data"])).number_format = "dd/mm/yyyy"
        ws.cell(i, 2, d.get("origem", ""))
        cat = (d.get("condicao") or "").capitalize()
        ws.cell(i, 3, rotulo(cat) if cat in CATEGORIAS else cat)
        ws.cell(i, 4, d.get("condicao_trabalho", ""))

    l0, col_total = _bloco_resumo(ws, 2, rotulo_bloco, historico)

    # Total de respostas = registros do InMeta + historico interno
    total = len(dias) + (sum(historico.values()) if historico else 0)

    # Caixas do padrao das reunioes, a esquerda da pizza
    _caixa(ws, "N3", "O5", f"TOTAL DE\n{total}\nRESPOSTAS", 12)
    _caixa(ws, "N14", "O16", rotulo_bloco, 10)
    for col in ("N", "O"):
        ws.column_dimensions[col].width = 16

    ws.add_chart(_pizza(ws, "Condição do Tempo", l0, col_total), "Q2")
    ws.freeze_panes = "A2"
    return ws, l0, col_total, total


def _fmt_br(iso: str) -> str:
    return datetime.date.fromisoformat(iso).strftime("%d/%m/%Y")


def build_tempo_report(
    dados: dict[str, Any],
    periodos: list[dict[str, str]] | None = None,
    n_meses: int = 3,
) -> bytes:
    """
    Monta o .xlsx no padrao das reunioes.

    dados: payload de /api/tempo (dias, historico, prioridade, ...)
    periodos: intervalos escolhidos na tela, ex.:
        [{"rotulo": "Período 1", "de": "2026-06-01", "ate": "2026-06-30"}, ...]
        Quando informados, cada um vira uma aba + pizza propria — reproduzindo
        as tres pizzas (total + dois periodos) usadas nas reunioes.
    n_meses: usado apenas quando nenhum periodo e informado (fallback).
    """
    dias = [d for d in dados.get("dias", []) if d.get("condicao") in CATEGORIAS
            or (d.get("condicao") or "").capitalize() in CATEGORIAS]
    # Mais recentes primeiro, como no modelo
    dias_ord = sorted(dias, key=lambda d: d["data"], reverse=True)

    wb = Workbook()
    resumo = wb.active
    resumo.title = "RESUMO"

    historico = {k.capitalize(): v for k, v in (dados.get("historico") or {}).items()}

    abas: list[tuple[str, str, Any, int, int, int]] = []

    # ── Aba TOTAL (inclui o historico interno pre-InMeta) ────────────────────
    _ws, l0, ct, tot = _escreve_aba_dados(
        wb, "DADOS TOTAL", dias_ord, "PERÍODO TOTAL OBRA", historico,
    )
    abas.append(("CONDIÇÃO DO TEMPO TOTAL OBRA", "DADOS TOTAL", _ws, l0, ct, tot))

    if periodos:
        # ── Uma aba por intervalo escolhido na tela ──────────────────────────
        for i, p in enumerate(periodos, start=1):
            de, ate = p.get("de", ""), p.get("ate", "")
            if not de or not ate:
                continue
            do_periodo = [d for d in dias_ord if de <= d["data"] <= ate]
            faixa = f"{_fmt_br(de)} a {_fmt_br(ate)}"
            rot = p.get("rotulo") or f"Período {i}"
            _ws, l0, ct, tot = _escreve_aba_dados(
                wb, f"DADOS {rot.upper()}"[:31], do_periodo,
                f"{rot.upper()} — {faixa}", None,
            )
            abas.append((f"CONDIÇÃO DO TEMPO {rot.upper()} | {faixa}",
                         rot, _ws, l0, ct, tot))
    else:
        # ── Fallback: ultimos meses com registro ─────────────────────────────
        meses_presentes = sorted({d["data"][:7] for d in dias_ord}, reverse=True)[:n_meses]
        for mes in meses_presentes:
            ano, m = mes.split("-")
            periodo = f"{_MESES[int(m) - 1].upper()} | {ano}"
            do_mes = [d for d in dias_ord if d["data"].startswith(mes)]
            nome_aba = f"DADOS {_MESES[int(m) - 1][:3].upper()} {ano}"
            _ws, l0, ct, tot = _escreve_aba_dados(
                wb, nome_aba, do_mes, f"PERÍODO {periodo}", None,
            )
            abas.append((f"CONDIÇÃO DO TEMPO {periodo}", nome_aba, _ws, l0, ct, tot))

    # ── RESUMO: cada pizza no seu bloco de colunas ───────────────────────────
    # Cada grafico tem 15 cm; espacar de 9 em 9 colunas evita a sobreposicao
    # que acontecia quando ficavam em colunas vizinhas.
    PASSO = 9
    for i, (titulo, _nome, ws_dados, l0, ct, tot) in enumerate(abas):
        col = 2 + i * PASSO
        letra = get_column_letter(col)

        c = resumo.cell(6, col, titulo)
        c.font = Font(name="Calibri", bold=True, size=11)
        c.alignment = Alignment(horizontal="left")

        _caixa(resumo, f"{letra}8", f"{get_column_letter(col + 1)}10",
               f"TOTAL DE\n{tot}\nRESPOSTAS", 12)
        resumo.column_dimensions[letra].width = 16
        resumo.column_dimensions[get_column_letter(col + 1)].width = 16

        resumo.add_chart(_pizza(ws_dados, "Condição do Tempo", l0, ct),
                         f"{get_column_letter(col + 2)}7")

    resumo.cell(2, 2, "DIÁRIO DO TEMPO — OBRAS").font = Font(bold=True, size=14)
    prioridade = " > ".join(dados.get("prioridade", []))
    resumo.cell(3, 2,
                f"Consolidado sem duplicar dias · prioridade: {prioridade}"
                ).font = Font(size=10, color="808080")
    resumo.cell(4, 2,
                f"Gerado em {datetime.datetime.now():%d/%m/%Y às %H:%M}"
                ).font = Font(size=10, color="808080")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
