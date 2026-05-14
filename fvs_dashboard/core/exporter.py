"""
core/exporter.py
================
Exportacao: Excel (bytes) + PDF resumo operacional (reportlab).
"""

from __future__ import annotations

import io
import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .business import STATUS_FINALIZADA, STATUS_EM_ANDAMENTO, STATUS_NAO_INICIADA

# ── Cores Excel ───────────────────────────────────────────────────────────────
C_AZUL       = "2F5597"
C_VERDE      = "375623"
C_AMARELO    = "7F6000"
C_VERMELHO   = "C00000"
C_CINZA      = "595959"
BG_VERDE     = "E2EFDA"
BG_AMARELO   = "FFEB9C"
BG_VERMELHO  = "FFC7CE"
BG_CINZA     = "F2F2F2"

STATUS_COLORS = {
    STATUS_FINALIZADA:   (C_VERDE,   BG_VERDE),
    STATUS_EM_ANDAMENTO: (C_AMARELO, BG_AMARELO),
    STATUS_NAO_INICIADA: (C_VERMELHO, BG_VERMELHO),
}

STATUS_LABEL = {
    STATUS_FINALIZADA:   "FINALIZADA",
    STATUS_EM_ANDAMENTO: "EM ANDAMENTO",
    STATUS_NAO_INICIADA: "NAO INICIADA",
}


def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)

def _font(color="000000", bold=False, size=9) -> Font:
    return Font(name="Calibri", color=color, bold=bold, size=size)

def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_w(ws, col: int, w: float):
    ws.column_dimensions[get_column_letter(col)].width = w


# ── Excel ─────────────────────────────────────────────────────────────────────

def _write_header(ws, text: str, ncols: int, row: int = 1) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row, 1, text)
    c.fill = _fill(C_AZUL)
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def _write_row(ws, row: int, values: list, fills: list | None = None,
               fonts: list | None = None, aligns: list | None = None) -> None:
    for j, v in enumerate(values, 1):
        c = ws.cell(row, j, v)
        c.border = _border()
        if fills:
            c.fill = _fill(fills[j - 1])
        if fonts:
            c.font = fonts[j - 1]
        if aligns:
            c.alignment = Alignment(horizontal=aligns[j - 1], wrap_text=True)


def export_excel(
    rows: list[dict],
    kpis: dict[str, Any],
    obra: str,
    include_finalizadas: bool = True,
) -> bytes:
    """
    Gera Excel com 4 abas e retorna bytes para st.download_button.
    """
    wb = Workbook()
    wb.remove(wb.active)
    today = datetime.date.today().strftime("%d/%m/%Y")

    if not include_finalizadas:
        rows = [r for r in rows if r["status"] != STATUS_FINALIZADA]

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resumo")
    ws_res.sheet_view.showGridLines = False
    _write_header(ws_res, f"RESUMO — {obra}   |   {today}", 4)

    data_res = [
        ("Atividades liberadas",         kpis["total_lib"]),
        ("Total FVS (associacoes)",       kpis["total_fvs"]),
        ("FINALIZADA",                    kpis["finalizada"]),
        ("EM ANDAMENTO",                  kpis["em_andamento"]),
        ("NAO INICIADA",                  kpis["nao_iniciada"]),
        ("Nao-conformidades abertas",     kpis["nc_total"]),
    ]
    for i, (label, val) in enumerate(data_res, 2):
        ws_res.cell(i, 1, label).font = _font(bold=True)
        ws_res.cell(i, 1).fill = _fill(BG_CINZA if i % 2 == 0 else "FFFFFF")
        ws_res.cell(i, 2, val).font  = _font()
        ws_res.cell(i, 2).alignment  = Alignment(horizontal="center")
    _set_w(ws_res, 1, 35)
    _set_w(ws_res, 2, 14)

    # ── Aba 2: Backlog ─────────────────────────────────────────────────────────
    ws_bl = wb.create_sheet("Backlog FVS")
    ws_bl.sheet_view.showGridLines = False
    hdrs = ["Pavimento", "WBS", "CF%", "Modelo FVS", "Local", "Status", "% Exec", "NC", "Data Insp."]
    widths = [22, 9, 6, 55, 30, 14, 8, 5, 12]
    _write_header(ws_bl, f"BACKLOG FVS — {obra}   |   {today}", len(hdrs))
    row = 2
    for j, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws_bl.cell(row, j, h)
        c.fill = _fill("1F3864"); c.font = _font("FFFFFF", True); c.alignment = Alignment(horizontal="center")
        _set_w(ws_bl, j, w)
    row += 1

    sorted_rows = sorted(rows, key=lambda r: (r.get("floor", ""), r.get("modelo", "")))
    for i, r in enumerate(sorted_rows):
        status = r["status"]
        fg, bg = STATUS_COLORS.get(status, (C_CINZA, BG_CINZA))
        alt_bg = BG_CINZA if i % 2 == 0 else "FFFFFF"
        vals = [
            r["floor"].split("|")[0].strip()[:25],
            r["wbs"],
            f"{r['cf_pct']:.0f}%",
            r["modelo"],
            r["local"],
            STATUS_LABEL.get(status, status),
            f"{r['pct_exec']}%" if r["pct_exec"] is not None else "—",
            str(r["nc"]) if r["nc"] else "",
            r["data_ins"],
        ]
        bg_list = [alt_bg] * 5 + [bg] + [alt_bg] * 3
        fn_list = [_font()] * 5 + [_font(fg, True)] + [_font()] * 3
        _write_row(ws_bl, row, vals, bg_list, fn_list,
                   ["left","center","center","left","left","center","center","center","center"])
        row += 1

    # ── Aba 3: Pendentes ────────────────────────────────────────────────────────
    ws_pend = wb.create_sheet("Pendentes")
    ws_pend.sheet_view.showGridLines = False
    pend = [r for r in rows if r["status"] == STATUS_NAO_INICIADA]
    hdrs_p = ["Pavimento", "WBS", "CF%", "Modelo FVS", "Local"]
    widths_p = [22, 9, 6, 55, 30]
    _write_header(ws_pend, f"FVS NAO INICIADAS — {obra}   |   {today}", len(hdrs_p))
    row = 2
    for j, (h, w) in enumerate(zip(hdrs_p, widths_p), 1):
        c = ws_pend.cell(row, j, h)
        c.fill = _fill("C00000"); c.font = _font("FFFFFF", True); c.alignment = Alignment(horizontal="center")
        _set_w(ws_pend, j, w)
    row += 1
    for i, r in enumerate(sorted(pend, key=lambda x: (x["modelo"], x["floor"]))):
        bg = "FFC7CE" if i % 2 == 0 else "FFD7D7"
        vals = [r["floor"].split("|")[0].strip()[:25], r["wbs"], f"{r['cf_pct']:.0f}%", r["modelo"], r["local"]]
        _write_row(ws_pend, row, vals, [bg]*5, [_font()]*5,
                   ["left","center","center","left","left"])
        row += 1
    ws_pend.cell(row + 1, 1, f"Total: {len(pend)} FVS nao iniciadas").font = _font(bold=True)

    # ── Aba 4: Por Modelo ─────────────────────────────────────────────────────
    ws_mod = wb.create_sheet("Por Modelo FVS")
    ws_mod.sheet_view.showGridLines = False
    from collections import defaultdict
    by_mod: dict[str, dict] = defaultdict(lambda: {STATUS_FINALIZADA: 0, STATUS_EM_ANDAMENTO: 0, STATUS_NAO_INICIADA: 0, "nc": 0})
    for r in rows:
        by_mod[r["modelo"]][r["status"]] += 1
        by_mod[r["modelo"]]["nc"] += r["nc"]
    hdrs_m = ["Modelo FVS", "Total", "Finalizada", "Em Andamento", "Nao Iniciada", "NC"]
    widths_m = [55, 8, 12, 14, 14, 8]
    _write_header(ws_mod, f"POR MODELO FVS — {obra}   |   {today}", len(hdrs_m))
    row = 2
    for j, (h, w) in enumerate(zip(hdrs_m, widths_m), 1):
        c = ws_mod.cell(row, j, h)
        c.fill = _fill("1F3864"); c.font = _font("FFFFFF", True); c.alignment = Alignment(horizontal="center")
        _set_w(ws_mod, j, w)
    row += 1
    for i, (mod_name, d) in enumerate(sorted(by_mod.items())):
        total = d[STATUS_FINALIZADA] + d[STATUS_EM_ANDAMENTO] + d[STATUS_NAO_INICIADA]
        bg = BG_CINZA if i % 2 == 0 else "FFFFFF"
        vals = [mod_name, total, d[STATUS_FINALIZADA], d[STATUS_EM_ANDAMENTO], d[STATUS_NAO_INICIADA], d["nc"]]
        bg_list = [bg, bg,
                   BG_VERDE    if d[STATUS_FINALIZADA]   else bg,
                   BG_AMARELO  if d[STATUS_EM_ANDAMENTO] else bg,
                   BG_VERMELHO if d[STATUS_NAO_INICIADA] else bg,
                   BG_VERMELHO if d["nc"] else bg]
        fn_list = [_font(bold=True)] + [_font()] * 5
        _write_row(ws_mod, row, vals, bg_list, fn_list,
                   ["left","center","center","center","center","center"])
        row += 1

    # ── Bytes ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ───────────────────────────────────────────────────────────────────────

def export_pdf(rows: list[dict], kpis: dict[str, Any], obra: str) -> bytes:
    """
    Gera PDF resumo operacional (2 paginas) com reportlab.
    Pagina 1: KPIs + grafico de pizza textual
    Pagina 2: tabela de FVS nao iniciadas
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    today  = datetime.date.today().strftime("%d/%m/%Y")

    # Estilos
    title_style  = ParagraphStyle("title",  fontSize=16, textColor=colors.HexColor("#2F5597"),
                                   fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=6)
    sub_style    = ParagraphStyle("sub",    fontSize=10, textColor=colors.grey,
                                   fontName="Helvetica", alignment=TA_CENTER, spaceAfter=12)
    section_style = ParagraphStyle("sect", fontSize=11, textColor=colors.HexColor("#1F3864"),
                                    fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=6)
    body_style   = styles["Normal"]

    story = []

    # ── Pagina 1: Cabecalho + KPIs ────────────────────────────────────────────
    story.append(Paragraph(f"RELATORIO FVS OPERACIONAL", title_style))
    story.append(Paragraph(f"{obra}   |   Gerado em {today}", sub_style))
    story.append(Spacer(1, 0.3*cm))

    # KPIs em tabela
    story.append(Paragraph("INDICADORES GERAIS", section_style))
    kpi_data = [
        ["Indicador", "Valor", "%"],
        ["Atividades liberadas para FVS", str(kpis["total_lib"]), "—"],
        ["Total de FVS (QAs)",           str(kpis["total_fvs"]), "100%"],
        ["FVS Finalizada",               str(kpis["finalizada"]),   f"{kpis['pct_finalizada']}%"],
        ["FVS Em Andamento",             str(kpis["em_andamento"]), f"{kpis['pct_em_andamento']}%"],
        ["FVS Nao Iniciada",             str(kpis["nao_iniciada"]), f"{kpis['pct_nao_iniciada']}%"],
        ["Nao-conformidades abertas",    str(kpis["nc_total"]),     "—"],
    ]
    kpi_table = Table(kpi_data, colWidths=[10*cm, 3*cm, 3*cm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#2F5597")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F2F2F2"), colors.white]),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN",        (1, 0), (-1, -1), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        # Cores por status
        ("BACKGROUND",   (0, 3), (-1, 3),  colors.HexColor("#E2EFDA")),  # Finalizada
        ("BACKGROUND",   (0, 4), (-1, 4),  colors.HexColor("#FFEB9C")),  # Em Andamento
        ("BACKGROUND",   (0, 5), (-1, 5),  colors.HexColor("#FFC7CE")),  # Nao Iniciada
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5*cm))

    # Distribuicao textual
    story.append(Paragraph("DISTRIBUICAO POR STATUS", section_style))
    total = kpis["total_fvs"] or 1
    bar_w = 40
    for status, count, pct in [
        ("Finalizada",   kpis["finalizada"],   kpis["pct_finalizada"]),
        ("Em Andamento", kpis["em_andamento"], kpis["pct_em_andamento"]),
        ("Nao Iniciada", kpis["nao_iniciada"], kpis["pct_nao_iniciada"]),
    ]:
        filled = int(bar_w * count / total)
        bar    = "=" * filled + "-" * (bar_w - filled)
        story.append(Paragraph(
            f"<font name='Courier'>{status:<14} [{bar}] {count:>4}  ({pct}%)</font>",
            body_style
        ))
    story.append(Spacer(1, 0.3*cm))

    # ── Pagina 2: FVS Nao Iniciadas ───────────────────────────────────────────
    story.append(PageBreak())
    story.append(Paragraph("FVS NAO INICIADAS — ACAO IMEDIATA", title_style))
    story.append(Paragraph(
        f"Estas FVS estao liberadas (execucao 100%) mas nao foram abertas no InMeta.",
        sub_style
    ))
    story.append(Spacer(1, 0.3*cm))

    pend = sorted(
        [r for r in rows if r["status"] == "NAO_INICIADA"],
        key=lambda r: (r["modelo"], r["floor"])
    )

    if pend:
        pend_data = [["Pavimento", "WBS", "Modelo FVS", "Local"]]
        for r in pend:
            pend_data.append([
                r["floor"].split("|")[0].strip()[:20],
                r["wbs"],
                r["modelo"][:50],
                r["local"][:30],
            ])
        pend_table = Table(pend_data, colWidths=[3.5*cm, 2*cm, 7.5*cm, 4.5*cm])
        pend_table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#C00000")),
            ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFC7CE"), colors.HexColor("#FFD7D7")]),
            ("GRID",         (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("WORDWRAP",     (2, 1), (2, -1),  True),
        ]))
        story.append(pend_table)
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f"<b>Total: {len(pend)} FVS nao iniciadas</b>",
            body_style
        ))
    else:
        story.append(Paragraph("Nenhuma FVS pendente. Todas estao iniciadas ou finalizadas.", body_style))

    doc.build(story)
    return buf.getvalue()
