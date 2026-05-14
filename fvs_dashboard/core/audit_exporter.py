"""
core/audit_exporter.py
======================
Exportadores de auditoria gerencial:
  - export_audit_pdf()  : PDF executivo 3 paginas com graficos Plotly embarcados
  - export_audit_excel(): Excel gerencial transposto (4 abas)

Fase 9: visual executivo — graficos Plotly → PNG (kaleido) → embutidos no PDF.
"""

from __future__ import annotations

import datetime
import io
from typing import Any

import pandas as pd

# ── Paleta corporativa ────────────────────────────────────────────────────────
C_AZUL      = "#2F5597"
C_AZUL_ESC  = "#1a2744"
C_AZUL_MED  = "#5b9bd5"
C_VERDE     = "#1e7e34"
C_AMARELO   = "#d39e00"
C_VERMELHO  = "#b21f2d"
C_CINZA_BG  = "#f5f7fa"
C_CINZA_LN  = "#dce3ed"
C_BRANCO    = "#FFFFFF"

# Hex sem # para openpyxl
OX_AZUL      = "2F5597"
OX_AZUL_ESC  = "1a2744"
OX_VERDE     = "1e7e34"
OX_AMARELO   = "d39e00"
OX_VERMELHO  = "b21f2d"
OX_CINZA_BG  = "f5f7fa"
OX_CINZA_LN  = "dce3ed"


# ═════════════════════════════════════════════════════════════════════════════
# GERAÇÃO DE GRAFICOS PLOTLY → PNG
# ═════════════════════════════════════════════════════════════════════════════

def _fig_to_png(fig, width: int, height: int) -> bytes:
    """Exporta figura Plotly para PNG de alta resolucao (kaleido, scale=2)."""
    return fig.to_image(format="png", width=width, height=height, scale=2)


def _chart_evolucao(monthly_insp: pd.DataFrame, obra_filter: str | None) -> bytes:
    """Grafico de linha: evolucao mensal de FVS por status."""
    import plotly.graph_objects as go

    df = monthly_insp.copy()
    if obra_filter:
        df = df[df["obra"] == obra_filter]

    months = sorted(df["date_month"].unique())
    labels = [m.strftime("%b/%y") for m in months]

    fin_v  = [int(df[df["date_month"] == m]["finalizada"].sum()) for m in months]
    em_v   = [int(df[df["date_month"] == m]["em_andamento"].sum()) for m in months]
    nc_v   = [int(df[df["date_month"] == m]["nc_total"].sum()) for m in months]

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=labels, y=fin_v, name="Finalizada",
        mode="lines+markers",
        line=dict(color=C_VERDE, width=3),
        marker=dict(size=8, symbol="circle"),
        fill="tozeroy", fillcolor="rgba(30,126,52,0.07)",
    ))
    fig.add_trace(go.Scatter(
        x=labels, y=em_v, name="Em Andamento",
        mode="lines+markers",
        line=dict(color=C_AMARELO, width=2.5),
        marker=dict(size=7),
    ))
    fig.add_trace(go.Scatter(
        x=labels, y=nc_v, name="NC Total",
        mode="lines+markers",
        line=dict(color=C_VERMELHO, width=2, dash="dot"),
        marker=dict(size=6, symbol="diamond"),
    ))
    fig.update_layout(
        title=dict(text="Evolução Mensal de FVS", font=dict(size=15, color=C_AZUL_ESC), x=0.02),
        paper_bgcolor=C_BRANCO, plot_bgcolor=C_CINZA_BG,
        height=340, margin=dict(l=50, r=30, t=45, b=55),
        legend=dict(orientation="h", y=-0.22, font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(tickfont=dict(size=10), gridcolor=C_CINZA_LN, showgrid=True, zeroline=False),
        yaxis=dict(tickfont=dict(size=10), gridcolor=C_CINZA_LN, showgrid=True, zeroline=False,
                   title=dict(text="Qtd FVS", font=dict(size=10))),
        font=dict(family="Arial, sans-serif"),
    )
    return _fig_to_png(fig, 860, 340)


def _chart_barras_empilhadas(monthly_insp: pd.DataFrame, obra_filter: str | None) -> bytes:
    """Barras empilhadas por mes: Finalizada + Em Andamento."""
    import plotly.graph_objects as go

    df = monthly_insp.copy()
    if obra_filter:
        df = df[df["obra"] == obra_filter]

    months = sorted(df["date_month"].unique())
    labels = [m.strftime("%b/%y") for m in months]
    fin_v  = [int(df[df["date_month"] == m]["finalizada"].sum()) for m in months]
    em_v   = [int(df[df["date_month"] == m]["em_andamento"].sum()) for m in months]

    fig = go.Figure()
    fig.add_trace(go.Bar(name="Finalizada",   x=labels, y=fin_v,
                         marker_color=C_VERDE,   opacity=0.88))
    fig.add_trace(go.Bar(name="Em Andamento", x=labels, y=em_v,
                         marker_color=C_AMARELO, opacity=0.88))
    fig.update_layout(
        barmode="stack",
        title=dict(text="Composição por Status (Mensal)", font=dict(size=14, color=C_AZUL_ESC), x=0.02),
        paper_bgcolor=C_BRANCO, plot_bgcolor=C_CINZA_BG,
        height=290, margin=dict(l=50, r=30, t=45, b=55),
        legend=dict(orientation="h", y=-0.25, font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(tickfont=dict(size=10), gridcolor=C_CINZA_LN),
        yaxis=dict(tickfont=dict(size=10), gridcolor=C_CINZA_LN, title=dict(text="Qtd", font=dict(size=10))),
        font=dict(family="Arial, sans-serif"),
    )
    return _fig_to_png(fig, 860, 290)


def _chart_pizza(latest_snap: pd.DataFrame) -> bytes:
    """Donut: distribuicao atual por status."""
    import plotly.graph_objects as go

    if latest_snap.empty:
        return b""
    fin  = int((latest_snap["status"] == "FINALIZADA").sum())
    em   = int((latest_snap["status"] == "EM_ANDAMENTO").sum())
    nao  = int((latest_snap["status"] == "NAO_INICIADA").sum())
    total = fin + em + nao

    fig = go.Figure(go.Pie(
        labels=["Finalizada", "Em Andamento", "Não Iniciada"],
        values=[fin, em, nao],
        hole=0.58,
        marker_colors=[C_VERDE, C_AMARELO, C_VERMELHO],
        textinfo="percent+label",
        textfont=dict(size=11),
        direction="clockwise",
    ))
    fig.update_layout(
        title=dict(text="Estado Atual das FVS", font=dict(size=14, color=C_AZUL_ESC), x=0.02),
        paper_bgcolor=C_BRANCO,
        height=300, margin=dict(l=20, r=20, t=45, b=20),
        showlegend=False,
        annotations=[dict(text=f"<b>{total}</b><br>FVS", x=0.5, y=0.5,
                          font=dict(size=15, color=C_AZUL_ESC), showarrow=False)],
        font=dict(family="Arial, sans-serif"),
    )
    return _fig_to_png(fig, 420, 300)


def _chart_comparativo(monthly_insp: pd.DataFrame) -> bytes:
    """Barras agrupadas: Cape Town vs Holmes por mes."""
    import plotly.graph_objects as go
    from fvs_dashboard.core.audit_engine import build_obra_comparison

    comp = build_obra_comparison(monthly_insp)
    if comp.empty:
        return b""

    # Limita aos ultimos 8 meses para nao poluir
    comp = comp.tail(8)
    labels = [m.strftime("%b/%y") for m in comp["date_month"]]

    fig = go.Figure()
    fig.add_trace(go.Bar(name="Cape Town", x=labels, y=comp["fin_ct"].tolist(),
                         marker_color=C_AZUL, opacity=0.88))
    fig.add_trace(go.Bar(name="Holmes",    x=labels, y=comp["fin_hm"].tolist(),
                         marker_color=C_AZUL_MED, opacity=0.88))
    fig.update_layout(
        barmode="group",
        title=dict(text="FVS Finalizadas — Cape Town vs Holmes", font=dict(size=14, color=C_AZUL_ESC), x=0.02),
        paper_bgcolor=C_BRANCO, plot_bgcolor=C_CINZA_BG,
        height=300, margin=dict(l=50, r=20, t=45, b=55),
        legend=dict(orientation="h", y=-0.28, font=dict(size=11), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(tickfont=dict(size=10), gridcolor=C_CINZA_LN),
        yaxis=dict(tickfont=dict(size=10), gridcolor=C_CINZA_LN),
        font=dict(family="Arial, sans-serif"),
    )
    return _fig_to_png(fig, 420, 300)


def _chart_aging(snap_latest: pd.DataFrame) -> bytes:
    """Barras horizontais: aging de backlog por faixa."""
    import plotly.graph_objects as go

    if snap_latest.empty:
        return b""
    pend = snap_latest[snap_latest["status"] != "FINALIZADA"]
    if pend.empty:
        return b""

    faixas = ["0-3d", "4-7d", "8-14d", ">14d"]
    vals   = [int((pend["faixa_aging"] == f).sum()) for f in faixas]
    cores  = [C_VERDE, C_AMARELO, "#e67e22", C_VERMELHO]

    fig = go.Figure(go.Bar(
        x=vals, y=faixas, orientation="h",
        marker_color=cores,
        text=[f"  {v}" for v in vals], textposition="outside",
        textfont=dict(size=11, color=C_AZUL_ESC),
    ))
    fig.update_layout(
        title=dict(text="Aging de Backlog (dias pendente)", font=dict(size=14, color=C_AZUL_ESC), x=0.02),
        paper_bgcolor=C_BRANCO, plot_bgcolor=C_CINZA_BG,
        height=240, margin=dict(l=60, r=60, t=45, b=20),
        xaxis=dict(tickfont=dict(size=10), gridcolor=C_CINZA_LN, showgrid=True),
        yaxis=dict(tickfont=dict(size=11), tickfont_color=C_AZUL_ESC),
        font=dict(family="Arial, sans-serif"),
    )
    return _fig_to_png(fig, 430, 240)


def _chart_top_modelos(snap_latest: pd.DataFrame) -> bytes:
    """Barras horizontais: top 8 modelos com mais FVS pendentes."""
    import plotly.graph_objects as go

    if snap_latest.empty:
        return b""
    pend = snap_latest[snap_latest["status"] != "FINALIZADA"]
    if pend.empty:
        return b""

    top = (
        pend.groupby("modelo")
        .agg(pendentes=("status", "count"), nc=("nc", "sum"))
        .sort_values("pendentes", ascending=True)
        .tail(8)
        .reset_index()
    )
    top["modelo_curto"] = top["modelo"].str.replace(r"FVS \d+\.\d+\.\d+ - ", "", regex=True).str[:32]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=top["modelo_curto"], x=top["pendentes"], orientation="h",
        name="Pendentes", marker_color=C_AZUL, opacity=0.88,
        text=top["pendentes"], textposition="auto",
    ))
    fig.add_trace(go.Bar(
        y=top["modelo_curto"], x=top["nc"], orientation="h",
        name="NC", marker_color=C_VERMELHO, opacity=0.65,
    ))
    fig.update_layout(
        barmode="overlay",
        title=dict(text="Top Gargalos por Modelo FVS", font=dict(size=14, color=C_AZUL_ESC), x=0.02),
        paper_bgcolor=C_BRANCO, plot_bgcolor=C_CINZA_BG,
        height=290, margin=dict(l=10, r=40, t=45, b=40),
        legend=dict(orientation="h", y=-0.22, font=dict(size=10), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(tickfont=dict(size=9), gridcolor=C_CINZA_LN),
        yaxis=dict(tickfont=dict(size=9)),
        font=dict(family="Arial, sans-serif"),
    )
    return _fig_to_png(fig, 430, 290)


def _chart_nc(monthly_insp: pd.DataFrame, obra_filter: str | None) -> bytes:
    """Linha: evolucao de NC (total, pendentes, tratadas)."""
    import plotly.graph_objects as go

    df = monthly_insp.copy()
    if obra_filter:
        df = df[df["obra"] == obra_filter]

    months = sorted(df["date_month"].unique())
    labels = [m.strftime("%b/%y") for m in months]
    nc_t   = [int(df[df["date_month"] == m]["nc_total"].sum()) for m in months]
    nc_p   = [int(df[df["date_month"] == m]["nc_pendentes"].sum()) for m in months]
    nc_tr  = [int(df[df["date_month"] == m]["nc_tratadas"].sum()) for m in months]

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=labels, y=nc_t,  name="NC Total",
                             mode="lines+markers", line=dict(color=C_VERMELHO, width=2.5),
                             marker=dict(size=7)))
    fig.add_trace(go.Scatter(x=labels, y=nc_p,  name="NC Pendentes",
                             mode="lines+markers", line=dict(color=C_AMARELO, width=2, dash="dash"),
                             marker=dict(size=6)))
    fig.add_trace(go.Scatter(x=labels, y=nc_tr, name="NC Tratadas",
                             mode="lines+markers", line=dict(color=C_VERDE, width=2, dash="dot"),
                             marker=dict(size=6)))
    fig.update_layout(
        title=dict(text="Evolução de Não-Conformidades", font=dict(size=14, color=C_AZUL_ESC), x=0.02),
        paper_bgcolor=C_BRANCO, plot_bgcolor=C_CINZA_BG,
        height=260, margin=dict(l=50, r=30, t=45, b=55),
        legend=dict(orientation="h", y=-0.28, font=dict(size=10), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(tickfont=dict(size=9), gridcolor=C_CINZA_LN),
        yaxis=dict(tickfont=dict(size=9), gridcolor=C_CINZA_LN),
        font=dict(family="Arial, sans-serif"),
    )
    return _fig_to_png(fig, 860, 260)


# ═════════════════════════════════════════════════════════════════════════════
# PDF EXECUTIVO
# ═════════════════════════════════════════════════════════════════════════════

def export_audit_pdf(
    monthly_insp: pd.DataFrame,
    kpis: dict[str, Any],
    obra: str,
    periodo_label: str,
    date_start: datetime.date,
    date_end: datetime.date,
    latest_snap: pd.DataFrame | None = None,
    hist_snap: pd.DataFrame | None = None,
) -> bytes:
    """
    Gera PDF executivo de auditoria — 3 paginas com graficos Plotly embarcados.

    Pagina 1: Visao Executiva — KPIs + grafico de evolucao temporal
    Pagina 2: Operacao — barras empilhadas + pizza + comparativo obras
    Pagina 3: Auditoria — aging + top gargalos + NC + alertas
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        HRFlowable, PageBreak, Image as RLImage, KeepTogether,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    # ── Constantes de layout ──────────────────────────────────────────────────
    PAGE_W, PAGE_H = A4  # 595 x 842 pt
    MARGIN         = 1.8 * cm
    CONTENT_W      = PAGE_W - 2 * MARGIN   # ~481 pt ≈ 17 cm
    HALF_W         = (CONTENT_W - 0.5*cm) / 2

    RL_AZUL  = colors.HexColor(C_AZUL)
    RL_ESC   = colors.HexColor(C_AZUL_ESC)
    RL_VERDE = colors.HexColor(C_VERDE)
    RL_AMAR  = colors.HexColor(C_AMARELO)
    RL_VERM  = colors.HexColor(C_VERMELHO)
    RL_BG    = colors.HexColor(C_CINZA_BG)
    RL_LN    = colors.HexColor(C_CINZA_LN)

    # ── Helpers de estilo ─────────────────────────────────────────────────────
    def _ps(name, **kw):
        kw.setdefault("fontName", "Helvetica")
        return ParagraphStyle(name, **kw)

    S_PAGE_TITLE = _ps("pt", fontSize=8, textColor=colors.HexColor("#9aa0b0"),
                        alignment=TA_RIGHT)
    S_SECTION    = _ps("sec", fontSize=10, textColor=RL_AZUL,
                        fontName="Helvetica-Bold", spaceBefore=6, spaceAfter=3)
    S_NORMAL     = _ps("nor", fontSize=9, textColor=RL_ESC, leading=13)
    S_SMALL      = _ps("sm",  fontSize=7.5, textColor=colors.HexColor("#8090a8"), leading=11)
    S_ALERT_TIT  = _ps("at", fontSize=9, textColor=RL_VERM, fontName="Helvetica-Bold", leading=12)
    S_CELL_L     = _ps("cl", fontSize=8, textColor=RL_ESC, alignment=TA_LEFT,  leading=11)
    S_CELL_R     = _ps("cr", fontSize=8, textColor=RL_ESC, alignment=TA_RIGHT, leading=11)
    S_CELL_H     = _ps("ch", fontSize=8, textColor=colors.white,
                        fontName="Helvetica-Bold", alignment=TA_CENTER, leading=11)

    def _hr():
        return HRFlowable(width="100%", thickness=1, color=RL_LN, spaceAfter=6, spaceBefore=2)

    def _png_image(png_bytes: bytes, width, height):
        if not png_bytes:
            return Spacer(1, height)
        return RLImage(io.BytesIO(png_bytes), width=width, height=height)

    # ── Construtor da capa de cada pagina ─────────────────────────────────────
    today_str  = datetime.date.today().strftime("%d/%m/%Y")
    report_ref = f"R21 EMPREENDIMENTOS  |  {obra}  |  {periodo_label}  |  {today_str}"

    def _page_header_band(title: str, subtitle: str = "") -> list:
        """Faixa de cabecalho escura no topo de cada pagina."""
        inner = [
            [Paragraph(title, _ps("ph", fontSize=14, textColor=colors.white,
                                  fontName="Helvetica-Bold"))],
        ]
        if subtitle:
            inner.append([Paragraph(subtitle, _ps("ps", fontSize=9,
                                                   textColor=colors.HexColor("#a0b8e0")))])
        tbl = Table(inner, colWidths=[CONTENT_W])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), RL_ESC),
            ("TOPPADDING",    (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ]))
        return [tbl, Spacer(1, 0.35*cm)]

    def _kpi_card(icon: str, value: str, label: str, sub: str, bar_color) -> Table:
        """Card de KPI corporativo com barra colorida lateral."""
        bar_tbl = Table([[""]], colWidths=[0.32*cm], rowHeights=[2.4*cm])
        bar_tbl.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), bar_color)]))

        content_tbl = Table([
            [Paragraph(f"<b>{value}</b>",
                       _ps("kv", fontSize=24, textColor=RL_ESC,
                            fontName="Helvetica-Bold"))],
            [Paragraph(label.upper(),
                       _ps("kl", fontSize=7.5, textColor=colors.HexColor("#6b7fa3"),
                            fontName="Helvetica-Bold", leading=10))],
            [Paragraph(sub, _ps("ks", fontSize=7.5,
                                 textColor=colors.HexColor("#8090a8"), leading=10))],
        ], colWidths=[CONTENT_W/3 - 0.7*cm])
        content_tbl.setStyle(TableStyle([
            ("TOPPADDING",    (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 1),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ]))

        outer = Table([[bar_tbl, content_tbl]], colWidths=[0.32*cm, CONTENT_W/3 - 0.44*cm])
        outer.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), colors.white),
            ("BOX",           (0,0), (-1,-1), 0.5, RL_LN),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 0),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ]))
        return outer

    # ── Gera graficos (PNG) antecipadamente ───────────────────────────────────
    obra_f = None if obra == "Todas as Obras" else obra

    snap_for_charts = latest_snap if latest_snap is not None and not latest_snap.empty \
                      else pd.DataFrame()
    if hist_snap is not None and not hist_snap.empty:
        lat_d = hist_snap["date_snapshot"].max()
        snap_lat = hist_snap[hist_snap["date_snapshot"] == lat_d]
        if obra_f:
            snap_lat = snap_lat[snap_lat["obra"] == obra_f]
    else:
        snap_lat = snap_for_charts

    png_evol  = _chart_evolucao(monthly_insp, obra_f)         if not monthly_insp.empty else b""
    png_bars  = _chart_barras_empilhadas(monthly_insp, obra_f) if not monthly_insp.empty else b""
    png_pizza = _chart_pizza(snap_for_charts)
    png_comp  = _chart_comparativo(monthly_insp)               if not monthly_insp.empty else b""
    png_aging = _chart_aging(snap_lat)
    png_top   = _chart_top_modelos(snap_lat)
    png_nc    = _chart_nc(monthly_insp, obra_f)                if not monthly_insp.empty else b""

    # ── Monta o documento ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
        title=f"Auditoria Gerencial — {obra}",
        author="R21 Empreendimentos",
    )

    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # PAGINA 1 — VISAO EXECUTIVA
    # ══════════════════════════════════════════════════════════════════════════

    story += _page_header_band(
        "AUDITORIA GERENCIAL — R21 EMPREENDIMENTOS",
        f"Obra: {obra}  ·  Período: {periodo_label}  ·  "
        f"{date_start.strftime('%d/%m/%Y')} a {date_end.strftime('%d/%m/%Y')}",
    )

    # 6 KPI cards (2 linhas de 3)
    total_insp = kpis.get("total_insp", 0)
    finalizada = kpis.get("finalizada", 0)
    em_and     = kpis.get("em_andamento", 0)
    snap_ni    = kpis.get("snap_nao_iniciada", 0)
    nc_total   = kpis.get("nc_total", 0)
    nc_pend    = kpis.get("snap_nc_pendentes", 0)
    pct_fin    = kpis.get("pct_finalizada", 0.0)
    criticas   = kpis.get("snap_criticas", 0)
    pct_em     = round(100 * em_and / total_insp, 1) if total_insp else 0.0

    row1 = [
        _kpi_card("📋", str(total_insp),  "FVS Inspecionadas",   "no periodo",         RL_AZUL),
        _kpi_card("✅", str(finalizada),  "FVS Finalizadas",     f"{pct_fin:.0f}%",    RL_VERDE),
        _kpi_card("🔄", str(em_and),      "Em Andamento",        f"{pct_em:.0f}%",     RL_AMAR),
    ]
    row2 = [
        _kpi_card("🔴", str(snap_ni),     "Nao Iniciadas",       "estado atual",       RL_VERM),
        _kpi_card("⚠️", str(nc_total),   "NC Total",            "no periodo",         RL_AMAR),
        _kpi_card("🔴", str(nc_pend),     "NC Pendentes",        "estado atual",       RL_VERM if nc_pend > 0 else RL_VERDE),
    ]
    gap = 0.25 * cm
    kpi_tbl1 = Table([row1], colWidths=[CONTENT_W/3 - gap/2]*3, hAlign="LEFT")
    kpi_tbl1.setStyle(TableStyle([("LEFTPADDING", (0,0),(-1,-1), gap/2),
                                   ("RIGHTPADDING",(0,0),(-1,-1), gap/2)]))
    kpi_tbl2 = Table([row2], colWidths=[CONTENT_W/3 - gap/2]*3, hAlign="LEFT")
    kpi_tbl2.setStyle(TableStyle([("LEFTPADDING", (0,0),(-1,-1), gap/2),
                                   ("RIGHTPADDING",(0,0),(-1,-1), gap/2)]))

    story.append(kpi_tbl1)
    story.append(Spacer(1, 0.3*cm))
    story.append(kpi_tbl2)
    story.append(Spacer(1, 0.5*cm))

    # Grafico principal: evolucao temporal (full width)
    story.append(Paragraph("Evolução Temporal de FVS", S_SECTION))
    story.append(_hr())
    story.append(_png_image(png_evol, CONTENT_W, 8.5*cm))
    story.append(Spacer(1, 0.15*cm))
    story.append(Paragraph(
        "Fonte: dataInspecao (InMeta) — único campo de data disponível na API. "
        "Representa data de criação/execução de cada inspeção FVS.",
        S_SMALL,
    ))

    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════════
    # PAGINA 2 — OPERAÇÃO
    # ══════════════════════════════════════════════════════════════════════════

    story += _page_header_band(
        "OPERAÇÃO — COMPOSIÇÃO E COMPARATIVO",
        report_ref,
    )

    # Barras empilhadas (full width)
    story.append(Paragraph("Composição Mensal por Status", S_SECTION))
    story.append(_hr())
    story.append(_png_image(png_bars, CONTENT_W, 7*cm))
    story.append(Spacer(1, 0.5*cm))

    # Pizza + Comparativo lado a lado
    story.append(Paragraph("Estado Atual e Comparativo entre Obras", S_SECTION))
    story.append(_hr())

    pizza_img = _png_image(png_pizza, HALF_W, 6.8*cm)
    comp_img  = _png_image(png_comp,  HALF_W, 6.8*cm)
    side_tbl  = Table([[pizza_img, comp_img]],
                       colWidths=[HALF_W + 0.2*cm, HALF_W + 0.2*cm])
    side_tbl.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
    ]))
    story.append(side_tbl)

    story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════════
    # PAGINA 3 — AUDITORIA E ALERTAS
    # ══════════════════════════════════════════════════════════════════════════

    story += _page_header_band(
        "AUDITORIA — AGING, GARGALOS E NÃO-CONFORMIDADES",
        report_ref,
    )

    # Aging + Top gargalos lado a lado
    story.append(Paragraph("Aging de Backlog e Top Gargalos", S_SECTION))
    story.append(_hr())

    aging_img = _png_image(png_aging, HALF_W, 5.6*cm)
    top_img   = _png_image(png_top,   HALF_W, 5.6*cm)
    side_tbl2 = Table([[aging_img, top_img]],
                       colWidths=[HALF_W + 0.2*cm, HALF_W + 0.2*cm])
    side_tbl2.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
    ]))
    story.append(side_tbl2)
    story.append(Spacer(1, 0.4*cm))

    # NC evolution (full width)
    story.append(Paragraph("Evolução de Não-Conformidades", S_SECTION))
    story.append(_hr())
    story.append(_png_image(png_nc, CONTENT_W, 6.2*cm))
    story.append(Spacer(1, 0.35*cm))

    # Alertas criticos
    story.append(Paragraph("Alertas Críticos", S_SECTION))
    story.append(_hr())

    if criticas > 0 and snap_for_charts is not None and not snap_for_charts.empty:
        crit_df = snap_for_charts[
            (snap_for_charts["status"] == "NAO_INICIADA") &
            (snap_for_charts["dias_pendente"] > 7)
        ].sort_values("dias_pendente", ascending=False).head(10)

        alert_hdr = [
            Paragraph("Obra",         S_CELL_H),
            Paragraph("Pavimento",    S_CELL_H),
            Paragraph("Modelo FVS",   S_CELL_H),
            Paragraph("Dias Pend.",   S_CELL_H),
            Paragraph("NC",           S_CELL_H),
        ]
        alert_rows = [alert_hdr]
        for _, r in crit_df.iterrows():
            modelo_curto = str(r.get("modelo", ""))
            modelo_curto = modelo_curto.replace("FVS ", "").split(" - ", 1)[-1][:35]
            bg = colors.HexColor("#fff5f5") if len(alert_rows) % 2 == 0 else colors.white
            alert_rows.append([
                Paragraph(str(r.get("obra", ""))[:18],        S_CELL_L),
                Paragraph(str(r.get("floor", ""))[:18],       S_CELL_L),
                Paragraph(modelo_curto,                        S_CELL_L),
                Paragraph(str(int(r.get("dias_pendente", 0))), S_CELL_R),
                Paragraph(str(int(r.get("nc", 0))),           S_CELL_R),
            ])

        alert_tbl = Table(alert_rows,
                           colWidths=[3.5*cm, 3.5*cm, 6.5*cm, 2*cm, 1.5*cm])
        alert_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), RL_VERM),
            ("ROWBACKGROUNDS",(0, 1), (-1,-1), [colors.white, colors.HexColor("#fff5f5")]),
            ("BOX",           (0, 0), (-1,-1), 0.5, RL_LN),
            ("INNERGRID",     (0, 0), (-1,-1), 0.3, RL_LN),
            ("TOPPADDING",    (0, 0), (-1,-1), 4),
            ("BOTTOMPADDING", (0, 0), (-1,-1), 4),
            ("LEFTPADDING",   (0, 1), (0,-1),  6),
        ]))
        story.append(alert_tbl)
    else:
        ok_tbl = Table([[
            Paragraph("✓  Nenhum alerta crítico — todas as FVS dentro do prazo.",
                      _ps("ok", fontSize=9, textColor=RL_VERDE, fontName="Helvetica-Bold"))
        ]], colWidths=[CONTENT_W])
        ok_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#d4edda")),
            ("BOX",           (0,0), (-1,-1), 0.5, RL_VERDE),
            ("TOPPADDING",    (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
            ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ]))
        story.append(ok_tbl)

    story.append(Spacer(1, 0.5*cm))

    # Rodape final
    nota_data = [[
        Paragraph(
            "Limitações: NC sem datas individuais de abertura/encerramento — endpoints inexistentes na API InMeta (HTTP 404). "
            "dataInspecao = único campo de data disponível (proxy de atividade). "
            f"Snapshots diários acumulados a partir de 14/05/2026. "
            f"Gerado em {today_str} — R21 Empreendimentos.",
            S_SMALL
        )
    ]]
    nota_tbl = Table(nota_data, colWidths=[CONTENT_W])
    nota_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#f0f4fa")),
        ("BOX",           (0,0), (-1,-1), 0.5, RL_LN),
        ("TOPPADDING",    (0,0), (-1,-1), 7),
        ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
    ]))
    story.append(nota_tbl)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════════════════════
# EXCEL GERENCIAL (inalterado)
# ═════════════════════════════════════════════════════════════════════════════

def export_audit_excel(
    monthly_insp: pd.DataFrame,
    kpis: dict[str, Any],
    obra: str,
    periodo_label: str,
    date_start: datetime.date,
    date_end: datetime.date,
    hist_snap: pd.DataFrame | None = None,
) -> bytes:
    """
    Gera Excel gerencial de auditoria (4 abas).
    Aba 1: Resumo — KPIs consolidados
    Aba 2: Mensal — indicadores nas linhas, meses nas colunas (transposto)
    Aba 3: Por Obra — Cape Town vs Holmes
    Aba 4: Aging — faixas por modelo FVS
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color="1a2744", size=10):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _border():
        s = Side(style="thin", color=OX_CINZA_LN)
        return Border(left=s, right=s, top=s, bottom=s)

    def _header_row(ws, row_num, values, bg=OX_AZUL):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=True, color="FFFFFF", size=10)
            c.alignment = _align("center")
            c.border    = _border()

    def _data_row(ws, row_num, values, bold=False, bg=None):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            if bg:
                c.fill = _fill(bg)
            c.font      = _font(bold=bold, size=9)
            c.alignment = _align("right" if col > 1 else "left")
            c.border    = _border()

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"

    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value     = "AUDITORIA GERENCIAL — R21 EMPREENDIMENTOS"
    c.fill      = _fill(OX_AZUL_ESC)
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.alignment = _align("center")
    ws1.row_dimensions[1].height = 28

    info = [
        ("Obra",        obra),
        ("Periodo",     periodo_label),
        ("Intervalo",   f"{date_start.strftime('%d/%m/%Y')} a {date_end.strftime('%d/%m/%Y')}"),
        ("Gerado em",   datetime.date.today().strftime("%d/%m/%Y")),
    ]
    for i, (k, v) in enumerate(info, 3):
        ws1.cell(i, 1, k).font = _font(bold=True, color=OX_AZUL)
        ws1.cell(i, 2, v).font = _font()
        ws1.row_dimensions[i].height = 16

    _header_row(ws1, 8, ["INDICADOR", "VALOR"], bg=OX_AZUL_ESC)
    ws1.row_dimensions[8].height = 18

    kpi_rows = [
        ("FVS Inspecionadas no Periodo",  kpis.get("total_insp", 0),           OX_AZUL),
        ("FVS Finalizadas",               kpis.get("finalizada", 0),            OX_VERDE),
        ("FVS Em Andamento",              kpis.get("em_andamento", 0),          OX_AMARELO),
        ("FVS Nao Iniciadas (atual)",     kpis.get("snap_nao_iniciada", 0),     OX_VERMELHO),
        ("FVS Criticas >7 dias",          kpis.get("snap_criticas", 0),         OX_VERMELHO),
        ("% FVS Finalizadas",             f"{kpis.get('pct_finalizada',0):.1f}%", OX_VERDE),
        ("NC Total no Periodo",           kpis.get("nc_total", 0),              OX_VERMELHO),
        ("NC Pendentes (atual)",          kpis.get("snap_nc_pendentes", 0),     OX_VERMELHO),
        ("NC Tratadas",                   kpis.get("nc_tratadas", 0),           OX_VERDE),
    ]
    for i, (label, val, cor) in enumerate(kpi_rows, 9):
        bg = OX_CINZA_BG if i % 2 == 0 else "FFFFFF"
        c1 = ws1.cell(i, 1, label)
        c1.font = _font(size=9); c1.fill = _fill(bg); c1.border = _border()
        c1.alignment = _align("left")
        c2 = ws1.cell(i, 2, val)
        c2.font = _font(bold=True, color=cor, size=10); c2.fill = _fill(bg)
        c2.border = _border(); c2.alignment = _align("right")

    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 18

    nota_row = len(kpi_rows) + 11
    ws1.merge_cells(f"A{nota_row}:D{nota_row}")
    nota_c = ws1.cell(nota_row, 1,
        "NOTA: dataInspecao = unica data disponivel na API InMeta. "
        "NC: sem datas individuais de abertura/encerramento (HTTP 404).")
    nota_c.font = _font(color="7f6000", size=8)
    nota_c.fill = _fill("fff3cd")
    nota_c.alignment = _align("left", wrap=True)
    ws1.row_dimensions[nota_row].height = 30

    # ── Aba 2: Mensal ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Mensal")

    if not monthly_insp.empty:
        months = sorted(monthly_insp["date_month"].unique())
        month_labels = [m.strftime("%b/%Y") for m in months]

        _header_row(ws2, 1, ["Indicador"] + month_labels + ["Total"])
        ws2.column_dimensions["A"].width = 32
        for i in range(2, len(months) + 3):
            ws2.column_dimensions[get_column_letter(i)].width = 12

        indicadores_men = [
            ("FVS Finalizadas",     "finalizada",    OX_VERDE),
            ("FVS Em Andamento",    "em_andamento",  OX_AMARELO),
            ("Total Inspecionadas", "total_insp",    OX_AZUL),
            ("NC Total",            "nc_total",       OX_VERMELHO),
            ("NC Pendentes",        "nc_pendentes",   OX_AMARELO),
            ("NC Tratadas",         "nc_tratadas",    OX_VERDE),
        ]
        for row_i, (label, col, cor) in enumerate(indicadores_men, 2):
            bg = OX_CINZA_BG if row_i % 2 == 0 else "FFFFFF"
            c_label = ws2.cell(row_i, 1, label)
            c_label.font = _font(bold=True, color=cor, size=9)
            c_label.fill = _fill(bg); c_label.border = _border()
            c_label.alignment = _align("left")
            total = 0
            for col_j, m in enumerate(months, 2):
                sub = monthly_insp[monthly_insp["date_month"] == m]
                val = int(sub[col].sum()) if col in sub.columns else 0
                total += val
                c = ws2.cell(row_i, col_j, val)
                c.font = _font(size=9); c.fill = _fill(bg)
                c.border = _border(); c.alignment = _align("right")
            c_tot = ws2.cell(row_i, len(months)+2, total)
            c_tot.font = _font(bold=True, color=cor, size=9)
            c_tot.fill = _fill(OX_CINZA_BG); c_tot.border = _border()
            c_tot.alignment = _align("right")
    else:
        ws2.cell(1, 1, "Sem dados de inspecoes no periodo.")

    # ── Aba 3: Por Obra ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Obra")
    from fvs_dashboard.core.audit_engine import build_obra_comparison
    comp = build_obra_comparison(monthly_insp) if not monthly_insp.empty else pd.DataFrame()

    _header_row(ws3, 1, [
        "Mes", "CT Finalizadas", "HM Finalizadas",
        "CT Em Andamento", "HM Em Andamento",
        "CT NC Total", "HM NC Total",
        "CT NC Pend.", "HM NC Pend.",
    ])
    for col_i in range(1, 10):
        ws3.column_dimensions[get_column_letter(col_i)].width = 16

    if not comp.empty:
        comp_filt = comp[
            (comp["date_month"] >= date_start) &
            (comp["date_month"] <= date_end)
        ]
        for row_i, (_, row) in enumerate(comp_filt.iterrows(), 2):
            bg = OX_CINZA_BG if row_i % 2 == 0 else "FFFFFF"
            _data_row(ws3, row_i, [
                row["date_month"].strftime("%b/%Y"),
                row["fin_ct"], row["fin_hm"],
                row["em_ct"],  row["em_hm"],
                row["nc_ct"],  row["nc_hm"],
                row["nc_pend_ct"], row["nc_pend_hm"],
            ], bg=bg)

    # ── Aba 4: Aging ──────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Aging")
    if hist_snap is not None and not hist_snap.empty:
        lat_d    = hist_snap["date_snapshot"].max()
        snap_lat = hist_snap[hist_snap["date_snapshot"] == lat_d]
        pend     = snap_lat[snap_lat["status"] != "FINALIZADA"]

        _header_row(ws4, 1, ["Modelo FVS", "Obra", "0-3d", "4-7d", "8-14d", ">14d", "Total Pend.", "NC"])
        ws4.column_dimensions["A"].width = 45
        ws4.column_dimensions["B"].width = 22
        for col_i in range(3, 9):
            ws4.column_dimensions[get_column_letter(col_i)].width = 12

        grp = (
            pend.groupby(["modelo", "obra"])
            .agg(
                f0_3  = ("faixa_aging", lambda s: (s == "0-3d").sum()),
                f4_7  = ("faixa_aging", lambda s: (s == "4-7d").sum()),
                f8_14 = ("faixa_aging", lambda s: (s == "8-14d").sum()),
                fgt14 = ("faixa_aging", lambda s: (s == ">14d").sum()),
                nc    = ("nc", "sum"),
            )
            .reset_index()
        )
        grp["total"] = grp["f0_3"] + grp["f4_7"] + grp["f8_14"] + grp["fgt14"]
        grp = grp.sort_values("total", ascending=False)

        for row_i, (_, row) in enumerate(grp.iterrows(), 2):
            bg = OX_CINZA_BG if row_i % 2 == 0 else "FFFFFF"
            _data_row(ws4, row_i, [
                row["modelo"], row["obra"],
                row["f0_3"], row["f4_7"], row["f8_14"], row["fgt14"],
                row["total"], row["nc"],
            ], bg=bg)
    else:
        ws4.cell(1, 1, "Sem historico de snapshots disponivel.")

    buf2 = io.BytesIO()
    wb.save(buf2)
    buf2.seek(0)
    return buf2.read()
